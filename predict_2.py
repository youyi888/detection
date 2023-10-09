import glob
import json
import math
import os
import time
import pandas as pd
from openpyxl import Workbook
from PyQt5.QtWidgets import QApplication, QFileDialog

import sys
import torch
import cv2
import numpy as np
from models.common import DetectMultiBackend
from utils.augmentations import letterbox
from utils.general import (Profile, check_img_size, non_max_suppression, scale_boxes)
from utils.torch_utils import select_device

os.environ['PATH'] = 'C:/tools/vips_dev/bin' + ';' + os.environ['PATH']

import pyvips

device = select_device('0')
seen, windows, dt = 0, [], (Profile(), Profile(), Profile())


def load_model(model_path, data_path):
    model = DetectMultiBackend(model_path, device, False, data_path, False)
    # stride, names, pt = model.stride, model.names, model.pt
    # print(stride, names, pt)
    return model


def prepare_image(image, img_size=640):
    img = letterbox(image.copy(), new_shape=img_size)[0]
    img = img[:, :, ::-1].transpose(2, 0, 1)  # BGR to RGB, to 3x416x416
    img = np.ascontiguousarray(img)
    return img, image


def load_image(im0, img_size, stride=32, auto=True):
    im = letterbox(im0, img_size, stride=stride, auto=auto)[0]  # padded resize
    im = im.transpose((2, 0, 1))[::-1]  # HWC to CHW, BGR to RGB
    im = np.ascontiguousarray(im)  # contiguous
    return im, im0


def get_abs_position(xywh, w, h):
    x1 = int((xywh[0] - xywh[2] / 2) * w)
    y1 = int((xywh[1] - xywh[3] / 2) * h)
    x2 = int((xywh[0] + xywh[2] / 2) * w)
    y2 = int((xywh[1] + xywh[3] / 2) * h)
    return x1, y1, x2, y2


def detect(model, img, imgsz=(1024, 1024)):
    stride, names, pt = model.stride, model.names, model.pt
    imgsz = check_img_size(imgsz, s=stride)
    model.warmup(imgsz=(1 if pt or model.triton else 1, 3, *imgsz))
    im, im0 = load_image(img, img_size=1024)

    with dt[0]:
        im = torch.from_numpy(im).to(model.device)
        im = im.half() if model.fp16 else im.float()  # uint8 to fp16/32
        im /= 255  # 0 - 255 to 0.0 - 1.0
        if len(im.shape) == 3:
            im = im[None]  # expand for batch dim

    with dt[1]:
        pred = model(im, augment=False, visualize=False)

    with dt[2]:
        pred = non_max_suppression(pred, 0.65, 0.35, None, False, max_det=1000)
    cls_result = []
    detection_result = []
    detection_score = []

    for i, det in enumerate(pred):
        if len(det):
            det[:, :4] = scale_boxes(im.shape[2:], det[:, :4], im0.shape).round()
            for *xyxy, conf, cls in reversed(det):
                print(conf)
                c = int(cls)
                _x0, _y0, _x1, _y1 = [int(x) for x in xyxy]
                cls_result.append(str(c))
                detection_result.append([str(c), _x0, _y0, _x1, _y1])
                detection_score.append(float(conf))
    return cls_result, detection_result, detection_score, img
def image_predict(image_path,out_path):
    image = cv2.imread(image_path)
    image = gamma_trans(image, 1)
    cls, bboxes, score, im = detect(md, image) 
    for bbox in bboxes:
        label, x1, y1, x2, y2 = bbox
        cv2.rectangle(image, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 2)
        text = str(md.names[int(label)]+format(score))
        cv2.putText(image, text, (int(x1), int(y1) - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
    cv2.imwrite(f"{out_path}/{parent_name}/3.jpg",image)
    cv2.imshow('Detection', image)
    cv2.waitKey(0)
def gamma_trans(img,gamma):#gamma大于1时图片变暗，小于1图片变亮
	#具体做法先归一化到1，然后gamma作为指数值求出新的像素值再还原
	gamma_table = [np.power(x/255.0,gamma)*255.0 for x in range(256)]
	gamma_table = np.round(np.array(gamma_table)).astype(np.uint8)
	#实现映射用的是Opencv的查表函数
	return cv2.LUT(img,gamma_table)

def colect():
    app = QApplication(sys.argv)
    file_dict = {}  # 存储文件夹路径的字典

    # 打开文件夹选择对话框，获取选择的文件夹路径列表
    dialog = QFileDialog()
    dialog.setFileMode(QFileDialog.Directory)
    dialog.setOption(QFileDialog.ShowDirsOnly)
    dialog.setOption(QFileDialog.DontUseNativeDialog)
    dialog.setOption(QFileDialog.ReadOnly)
    dialog.setOption(QFileDialog.DontUseCustomDirectoryIcons)

    if dialog.exec():
        folders = dialog.selectedFiles()

    # 将文件夹路径存储到字典中
    for folder in folders:
        folder_name = dialog.directory().relativeFilePath(folder)  # 获取文件夹名称
        file_dict[folder_name] = folder 
        return(folder)




if __name__ == '__main__':
    
    # 创建一个空的数据框
    combined_df = pd.DataFrame()
    md = load_model('best.pt', 'mydata2.yaml')
    out_path = f'C:/Users/Administrator/Contacts/Desktop/detection/1'
    # input_image = 'E:/mark10.7/1-2/tile/focus0/0/039x050.jpg'
    # input_path = 'E:/mark10.7'
    input_path=colect()
    for i in range(5):
        print(input_path)
        # image_predict(input_image,out_path)
        if not os.path.exists(out_path):
            os.makedirs(out_path)
        # 创建一个空的 Excel 工作簿
        starttime = time.time()
        workbook = Workbook()
        sheet = workbook.active
        # 设置表格列名
        sheet.cell(row=1, column=1, value='imagefolder')
        sheet.cell(row=1, column=2, value='imagepath')
        sheet.cell(row=1, column=3, value='label')
        sheet.cell(row=1, column=4, value='x1')
        sheet.cell(row=1, column=5, value='y1')
        sheet.cell(row=1, column=6, value='x2')
        sheet.cell(row=1, column=7, value='y2')
        for root, dirs, files in os.walk(input_path):
            starttime = time.time()
            for dir_name in dirs:
                if dir_name == "0":
                    dir_path = os.path.join(root, dir_name)
                    # 遍历0号文件夹内的图片文件
                    for file_name in os.listdir(dir_path):
                        if file_name.endswith('.jpg'):
                            file_path = os.path.join(dir_path, file_name)
                            # 在此处执行对jpg文件的处理操作
                            # print(file_path)  # 示例：打印文件路径
        # files = os.listdir(input_path)

        # for file in files:
        #     if not file.endswith('.jpg'):
        #         continue                            # 获取上一级目录路径
                            parent_path_1 = os.path.dirname(dir_path)
                            parent_path_2 = os.path.dirname(parent_path_1)
                            parent_path_3 = os.path.dirname(parent_path_2)
                            # 获取上一级目录名称
                            parent_name = os.path.basename(parent_path_3)

                            # img_path = os.path.join(input_path, file)
                            image = cv2.imread(file_path)
                            image = gamma_trans(image, 0.75)
                            cls, bboxes, score, im = detect(md, image)
                    
                            for bbox in bboxes:
                                label, x1, y1, x2, y2 = bbox
                                t = {"image_folder":parent_name,
                                    "imagepath":file_path,
                                    'label': md.names[int(label)],
                                    'x1': x1,
                                    'y1': y1,
                                    'x2': x2,
                                    'y2': y2
                                }
                                # print(t)
                                sheet.append([t['image_folder'],t['imagepath'],t['label'], t['x1'], t['y1'], t['x2'], t['y2']])
                                        # 在图像上绘制预测框
                                cv2.rectangle(image, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 2)
                                text = str(t['label']+format(score))
                                cv2.putText(image, text, (int(x1), int(y1) - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                            # 显示图像和预测框
                            if bboxes:
                            # cv2.imshow('Detection', image)
                            # cv2.waitKey(0)


                                if not os.path.exists(f"{out_path}/{parent_name}/{parent_name}_{i}"):
                                    os.makedirs(f"{out_path}/{parent_name}/{parent_name}_{i}")
                                cv2.imwrite(f"{out_path}/{parent_name}/{parent_name}_{i}/{file_name}",image)
                        if not os.path.exists(f"{out_path}/{parent_name}/{parent_name}_{i}"):
                            os.makedirs(f"{out_path}/{parent_name}/{parent_name}_{i}")
                        excel_file = os.path.join(f"{out_path}/{parent_name}/{parent_name}_{i}", f'{parent_name}_{i}.xlsx')
                        workbook.save(excel_file)
                    df = pd.read_excel(excel_file)

                    # 在单个表格开头插入一个带有表格名称的行
                    table_name_row = pd.Series([f'Table: detection_results_{i}'])
                    df.insert(loc=0, column='Table Name', value=table_name_row)

                    # 将单个表格添加到合并的数据框中
                    combined_df = pd.concat([combined_df, df], ignore_index=True)
                    endtime = time.time() - starttime
                    print("总共时长",endtime)
        # 保存合并后的数据框为 Excel 文件
    combined_df.to_excel(f'{out_path}/{parent_name}/{parent_name}.xlsx', index=False)       


